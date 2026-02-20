"""Auditor router for accessing inventory and activity logs."""

import io
import math
from collections import defaultdict
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import Employee, Customer, Product, Order, ReturnRequest
from app.models.inventory_log import InventoryLog, InventoryActionType, RelatedEntityType
from app.models.activity_log import ActivityLog, ActivityType
from app.models.inventory_transaction import InventoryTransaction
from app.models.order_item import OrderItem
from app.schemas.inventory_log import (
    InventoryLogResponse,
    InventoryLogListResponse,
)
from app.schemas.activity_log import (
    ActivityLogResponse,
    ActivityLogListResponse,
)
from app.utils.deps import require_auditor

router = APIRouter(prefix="/auditor", tags=["Auditor"])


def get_actor_info(log: InventoryLog, db: Session) -> tuple:
    """Get actor name, email, and type for an inventory log."""
    if log.actor_customer_id:
        customer = db.query(Customer).filter(Customer.id == log.actor_customer_id).first()
        if customer:
            return customer.full_name, customer.email, "customer"
    elif log.actor_employee_id:
        employee = db.query(Employee).filter(Employee.id == log.actor_employee_id).first()
        if employee:
            return employee.full_name, employee.email, "employee"
    return None, None, None


def get_user_info(log: ActivityLog, db: Session) -> tuple:
    """Get user name, email, and type for an activity log."""
    if log.customer_id:
        customer = db.query(Customer).filter(Customer.id == log.customer_id).first()
        if customer:
            return customer.full_name, customer.email, "customer"
    elif log.employee_id:
        employee = db.query(Employee).filter(Employee.id == log.employee_id).first()
        if employee:
            return employee.full_name, employee.email, "employee"
    return None, None, None


def get_related_entity_summary(log: InventoryLog, db: Session) -> Optional[str]:
    """Get a summary of the related entity for an inventory log."""
    if not log.related_entity_type or not log.related_entity_id:
        return None

    if log.related_entity_type == RelatedEntityType.ORDER:
        order = db.query(Order).filter(Order.id == log.related_entity_id).first()
        if order:
            return f"Order #{str(order.id)[:8]} - Rs. {order.total_amount:.2f}"
    elif log.related_entity_type == RelatedEntityType.PRODUCT:
        product = db.query(Product).filter(Product.id == log.related_entity_id).first()
        if product:
            return f"{product.name} ({product.brand})"
    elif log.related_entity_type == RelatedEntityType.RETURN_REQUEST:
        return_req = db.query(ReturnRequest).filter(ReturnRequest.id == log.related_entity_id).first()
        if return_req:
            return f"Return #{str(return_req.id)[:8]} - {return_req.reason}"

    return None


def inventory_log_to_response(log: InventoryLog, db: Session) -> InventoryLogResponse:
    """Convert InventoryLog model to InventoryLogResponse with related data."""
    actor_name, actor_email, actor_type = get_actor_info(log, db)
    related_entity_summary = get_related_entity_summary(log, db)

    return InventoryLogResponse(
        id=log.id,
        actor_customer_id=log.actor_customer_id,
        actor_employee_id=log.actor_employee_id,
        actor_name=actor_name,
        actor_email=actor_email,
        actor_type=actor_type,
        action_type=log.action_type,
        description=log.description,
        related_entity_type=log.related_entity_type,
        related_entity_id=log.related_entity_id,
        related_entity_summary=related_entity_summary,
        created_at=log.created_at
    )


def activity_log_to_response(log: ActivityLog, db: Session) -> ActivityLogResponse:
    """Convert ActivityLog model to ActivityLogResponse with related data."""
    user_name, user_email, user_type = get_user_info(log, db)

    return ActivityLogResponse(
        id=log.id,
        customer_id=log.customer_id,
        employee_id=log.employee_id,
        user_name=user_name,
        user_email=user_email,
        user_type=user_type,
        activity_type=log.activity_type,
        description=log.description,
        ip_address=log.ip_address,
        user_agent=log.user_agent,
        created_at=log.created_at
    )


@router.get("/inventory-logs", response_model=InventoryLogListResponse)
async def get_inventory_logs(
    action_type: Optional[InventoryActionType] = Query(None, description="Filter by action type"),
    related_entity_type: Optional[RelatedEntityType] = Query(None, description="Filter by related entity type"),
    actor_type: Optional[str] = Query(None, description="Filter by actor type (customer or employee)"),
    from_date: Optional[datetime] = Query(None, description="Filter from this date"),
    to_date: Optional[datetime] = Query(None, description="Filter until this date"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Get paginated inventory logs. Auditor or Admin only.

    Returns inventory-related events including:
    - Order placement and status changes
    - Product creation, updates, and deletions
    - Stock adjustments
    - Return request lifecycle changes
    """
    query = db.query(InventoryLog)

    # Apply filters
    if action_type:
        query = query.filter(InventoryLog.action_type == action_type)
    if related_entity_type:
        query = query.filter(InventoryLog.related_entity_type == related_entity_type)
    if actor_type == "customer":
        query = query.filter(InventoryLog.actor_customer_id.isnot(None))
    elif actor_type == "employee":
        query = query.filter(InventoryLog.actor_employee_id.isnot(None))
    if from_date:
        query = query.filter(InventoryLog.created_at >= from_date)
    if to_date:
        query = query.filter(InventoryLog.created_at <= to_date)

    # Get total count
    total = query.count()

    # Apply pagination and sorting
    offset = (page - 1) * page_size
    logs = query.order_by(InventoryLog.created_at.desc()).offset(offset).limit(page_size).all()

    total_pages = math.ceil(total / page_size) if total > 0 else 1

    return InventoryLogListResponse(
        items=[inventory_log_to_response(log, db) for log in logs],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/activity-logs", response_model=ActivityLogListResponse)
async def get_activity_logs(
    activity_type: Optional[ActivityType] = Query(None, description="Filter by activity type"),
    user_type: Optional[str] = Query(None, description="Filter by user type (customer or employee)"),
    from_date: Optional[datetime] = Query(None, description="Filter from this date"),
    to_date: Optional[datetime] = Query(None, description="Filter until this date"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Get paginated activity logs. Auditor or Admin only.

    Returns user account-related activities including:
    - User registration
    - Login and logout
    - Password changes
    - Profile updates
    - Role changes
    - Account deletion
    """
    query = db.query(ActivityLog)

    # Apply filters
    if activity_type:
        query = query.filter(ActivityLog.activity_type == activity_type)
    if user_type == "customer":
        query = query.filter(ActivityLog.customer_id.isnot(None))
    elif user_type == "employee":
        query = query.filter(ActivityLog.employee_id.isnot(None))
    if from_date:
        query = query.filter(ActivityLog.created_at >= from_date)
    if to_date:
        query = query.filter(ActivityLog.created_at <= to_date)

    # Get total count
    total = query.count()

    # Apply pagination and sorting
    offset = (page - 1) * page_size
    logs = query.order_by(ActivityLog.created_at.desc()).offset(offset).limit(page_size).all()

    total_pages = math.ceil(total / page_size) if total > 0 else 1

    return ActivityLogListResponse(
        items=[activity_log_to_response(log, db) for log in logs],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/inventory-logs/export")
async def export_inventory_logs(
    action_type: Optional[InventoryActionType] = Query(None),
    related_entity_type: Optional[RelatedEntityType] = Query(None),
    actor_type: Optional[str] = Query(None),
    from_date: Optional[datetime] = Query(None),
    to_date: Optional[datetime] = Query(None),
    search: Optional[str] = Query(None, description="Search by actor name, email, or description"),
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Export all matching inventory logs as Excel. Auditor or Admin only."""
    query = db.query(InventoryLog)

    if action_type:
        query = query.filter(InventoryLog.action_type == action_type)
    if related_entity_type:
        query = query.filter(InventoryLog.related_entity_type == related_entity_type)
    if actor_type == "customer":
        query = query.filter(InventoryLog.actor_customer_id.isnot(None))
    elif actor_type == "employee":
        query = query.filter(InventoryLog.actor_employee_id.isnot(None))
    if from_date:
        query = query.filter(InventoryLog.created_at >= from_date)
    if to_date:
        query = query.filter(InventoryLog.created_at <= to_date)

    logs = query.order_by(InventoryLog.created_at.desc()).all()
    responses = [inventory_log_to_response(log, db) for log in logs]

    # Apply search filter (client-side on the enriched data)
    if search:
        search_lower = search.lower()
        responses = [
            r for r in responses
            if (r.actor_name and search_lower in r.actor_name.lower())
            or (r.actor_email and search_lower in r.actor_email.lower())
            or (search_lower in r.description.lower())
        ]

    wb = _build_inventory_logs_workbook(responses)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_logs_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/activity-logs/export")
async def export_activity_logs(
    activity_type: Optional[ActivityType] = Query(None),
    user_type: Optional[str] = Query(None),
    from_date: Optional[datetime] = Query(None),
    to_date: Optional[datetime] = Query(None),
    search: Optional[str] = Query(None, description="Search by user name, email, or description"),
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Export all matching activity logs as Excel. Auditor or Admin only."""
    query = db.query(ActivityLog)

    if activity_type:
        query = query.filter(ActivityLog.activity_type == activity_type)
    if user_type == "customer":
        query = query.filter(ActivityLog.customer_id.isnot(None))
    elif user_type == "employee":
        query = query.filter(ActivityLog.employee_id.isnot(None))
    if from_date:
        query = query.filter(ActivityLog.created_at >= from_date)
    if to_date:
        query = query.filter(ActivityLog.created_at <= to_date)

    logs = query.order_by(ActivityLog.created_at.desc()).all()
    responses = [activity_log_to_response(log, db) for log in logs]

    # Apply search filter (client-side on the enriched data)
    if search:
        search_lower = search.lower()
        responses = [
            r for r in responses
            if (r.user_name and search_lower in r.user_name.lower())
            or (r.user_email and search_lower in r.user_email.lower())
            or (search_lower in r.description.lower())
        ]

    wb = _build_activity_logs_workbook(responses)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"activity_logs_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Report Download Helpers ──────────────────────────────────────────────────

def _fmt_dt(dt) -> str:
    """Format naive Sri Lanka datetime for Excel display."""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _build_inventory_logs_workbook(responses) -> Workbook:
    """Build an Excel workbook from a list of InventoryLogResponse objects."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Audit Logs"
    headers = [
        "Timestamp", "Action Type", "Actor Name", "Actor Email", "Actor Type",
        "Description", "Related Entity Type", "Related Entity Summary",
    ]
    _style_header_row(ws, headers)
    for r in responses:
        action = r.action_type.value if hasattr(r.action_type, 'value') else str(r.action_type)
        entity_type = r.related_entity_type.value if r.related_entity_type and hasattr(r.related_entity_type, 'value') else (str(r.related_entity_type) if r.related_entity_type else "")
        ws.append([
            _fmt_dt(r.created_at),
            action,
            r.actor_name or "",
            r.actor_email or "",
            r.actor_type or "",
            r.description,
            entity_type,
            r.related_entity_summary or "",
        ])
    return wb


def _build_activity_logs_workbook(responses) -> Workbook:
    """Build an Excel workbook from a list of ActivityLogResponse objects."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Audit Logs"
    headers = [
        "Timestamp", "Activity Type", "User Name", "User Email", "User Type",
        "Description", "IP Address",
    ]
    _style_header_row(ws, headers)
    for r in responses:
        activity = r.activity_type.value if hasattr(r.activity_type, 'value') else str(r.activity_type)
        ws.append([
            _fmt_dt(r.created_at),
            activity,
            r.user_name or "",
            r.user_email or "",
            r.user_type or "",
            r.description,
            r.ip_address or "",
        ])
    return wb


def _style_header_row(ws, headers: list):
    """Bold white text on dark teal background for the header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)


# ─── Report Download Endpoints ────────────────────────────────────────────────

@router.get("/reports/monthly-transactions")
async def download_monthly_transaction_report(
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Download all orders as a two-sheet Excel report. Auditor or Admin only.

    Sheet 1 - Orders Summary: one row per order with customer/product info.
    Sheet 2 - Monthly Summary: aggregated totals grouped by month.
    """
    orders = (
        db.query(Order)
        .options(
            joinedload(Order.customer),
            joinedload(Order.items).joinedload(OrderItem.product),
        )
        .order_by(Order.created_at.asc())
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Orders Summary ──
    ws1 = wb.active
    ws1.title = "Orders Summary"
    headers1 = [
        "Month", "Order ID", "Customer Name", "Customer Email",
        "Sales Channel", "Status", "Payment Status", "Delivery Method",
        "Total Amount (LKR)", "Items Count", "Created At",
    ]
    _style_header_row(ws1, headers1)

    for order in orders:
        month = order.created_at.strftime("%Y-%m") if order.created_at else ""
        order_id_short = str(order.id)[:8].upper()

        if order.customer_id and order.customer:
            customer_name = order.customer.full_name
            customer_email = order.customer.email
        else:
            customer_name = order.offline_customer_name or "(walk-in)"
            customer_email = "(offline sale)"

        items_count = sum(item.quantity for item in order.items)

        ws1.append([
            month,
            order_id_short,
            customer_name,
            customer_email,
            order.sales_channel.value if order.sales_channel else "",
            order.status.value if order.status else "",
            order.payment_status.value if hasattr(order, "payment_status") and order.payment_status else "",
            order.delivery_method.value if order.delivery_method else "",
            float(order.total_amount) if order.total_amount is not None else 0.0,
            items_count,
            _fmt_dt(order.created_at),
        ])

    # ── Sheet 2: Monthly Summary ──
    ws2 = wb.create_sheet(title="Monthly Summary")
    headers2 = [
        "Month", "Total Orders", "Total Revenue (LKR)",
        "Online Orders", "Offline Orders", "Delivered", "Cancelled",
    ]
    _style_header_row(ws2, headers2)

    monthly: dict = defaultdict(lambda: {
        "total_orders": 0,
        "total_revenue": 0.0,
        "online": 0,
        "offline": 0,
        "delivered": 0,
        "cancelled": 0,
    })

    for order in orders:
        month = order.created_at.strftime("%Y-%m") if order.created_at else "Unknown"
        m = monthly[month]
        m["total_orders"] += 1
        status_val = order.status.value if order.status else ""
        channel_val = order.sales_channel.value if order.sales_channel else ""
        if status_val != "cancelled":
            m["total_revenue"] += float(order.total_amount) if order.total_amount else 0.0
        if channel_val == "online":
            m["online"] += 1
        elif channel_val == "offline":
            m["offline"] += 1
        if status_val == "delivered":
            m["delivered"] += 1
        elif status_val == "cancelled":
            m["cancelled"] += 1

    for month_key in sorted(monthly.keys()):
        m = monthly[month_key]
        ws2.append([
            month_key,
            m["total_orders"],
            round(m["total_revenue"], 2),
            m["online"],
            m["offline"],
            m["delivered"],
            m["cancelled"],
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"monthly_transactions_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/inventory-audit")
async def download_inventory_audit_report(
    current_employee: Employee = Depends(require_auditor),
    db: Session = Depends(get_db)
):
    """Download inventory transactions + current stock levels as a two-sheet Excel. Auditor or Admin only.

    Sheet 1 - Transactions: full history of every inventory change.
    Sheet 2 - Current Stock Levels: snapshot of all active products today.
    """
    LOW_STOCK_THRESHOLD = 3

    transactions = (
        db.query(InventoryTransaction)
        .options(
            joinedload(InventoryTransaction.product),
            joinedload(InventoryTransaction.employee),
        )
        .order_by(InventoryTransaction.created_at.asc())
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Transactions ──
    ws1 = wb.active
    ws1.title = "Transactions"
    headers1 = [
        "Date", "Product Name", "Brand", "Category",
        "Transaction Type", "Quantity Change", "Qty Before", "Qty After",
        "Reason", "Done By (Name)", "Done By (Email)",
    ]
    _style_header_row(ws1, headers1)

    for txn in transactions:
        product = txn.product
        employee = txn.employee
        ws1.append([
            _fmt_dt(txn.created_at),
            product.name if product else "(deleted)",
            product.brand if product else "",
            product.category if product else "",
            txn.transaction_type.value if txn.transaction_type else "",
            txn.quantity_change,
            txn.quantity_before,
            txn.quantity_after,
            txn.reason or "",
            employee.full_name if employee else "(system)",
            employee.email if employee else "",
        ])

    # ── Sheet 2: Current Stock Levels ──
    ws2 = wb.create_sheet(title="Current Stock Levels")
    headers2 = [
        "Product Name", "Brand", "Category",
        "Unit Price (LKR)", "Current Stock", "Stock Value (LKR)", "Status",
    ]
    _style_header_row(ws2, headers2)

    products = (
        db.query(Product)
        .filter(Product.is_active == True)
        .order_by(Product.name)
        .all()
    )

    for product in products:
        qty = product.quantity_available
        if qty == 0:
            status = "Out of Stock"
        elif qty <= LOW_STOCK_THRESHOLD:
            status = "Low Stock"
        else:
            status = "In Stock"

        unit_price = float(product.price) if product.price is not None else 0.0
        stock_value = round(unit_price * qty, 2)

        ws2.append([
            product.name,
            product.brand,
            product.category,
            unit_price,
            qty,
            stock_value,
            status,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_audit_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
